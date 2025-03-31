import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def json_to_excel(json_file_path, excel_file_path=None):
    """
    将template.json文件转换为Excel文件
    
    Args:
        json_file_path: JSON文件路径
        excel_file_path: Excel文件路径，默认为JSON文件同目录下的同名Excel文件
    """
    # 如果未指定Excel文件路径，则使用默认路径
    if excel_file_path is None:
        excel_file_path = os.path.splitext(json_file_path)[0] + '.xlsx'
    
    # 加载JSON数据
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            templates = json.load(f)
    except Exception as e:
        print(f"加载JSON文件失败: {e}")
        return False
    
    # 创建Excel工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模板数据"
    
    # 设置表头
    headers = ["模板名称", "优先级", "示例", "领域(origin)", "意图(origin)", "槽位(origin)", 
              "领域(last)", "意图(last)", "槽位(last)", "模板内容"]
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    row_idx = 2
    for template in templates:
        # 基本信息
        ws.cell(row=row_idx, column=1, value=template.get("name", ""))
        ws.cell(row=row_idx, column=2, value=template.get("priority", ""))
        
        # 示例（可能有多个，用换行符连接）
        examples = "\n".join(template.get("examples", []))
        ws.cell(row=row_idx, column=3, value=examples)
        
        # 条件信息
        conditions = template.get("conditions", {})
        
        # origin_slot信息
        origin_slot = conditions.get("origin_slot", {})
        ws.cell(row=row_idx, column=4, value="\n".join(origin_slot.get("domain", [])))
        ws.cell(row=row_idx, column=5, value="\n".join(origin_slot.get("intent", [])))
        
        # origin_slot的槽位（可能有多个，每个是一个字典）
        origin_slots_text = []
        for slot_dict in origin_slot.get("slots", []):
            for key, value in slot_dict.items():
                origin_slots_text.append(f"{key}: {value}")
        ws.cell(row=row_idx, column=6, value="\n".join(origin_slots_text))
        
        # last_slot信息
        last_slot = conditions.get("last_slot", {})
        ws.cell(row=row_idx, column=7, value="\n".join(last_slot.get("domain", [])))
        ws.cell(row=row_idx, column=8, value="\n".join(last_slot.get("intent", [])))
        
        # last_slot的槽位
        last_slots_text = []
        for slot_dict in last_slot.get("slots", []):
            for key, value in slot_dict.items():
                last_slots_text.append(f"{key}: {value}")
        ws.cell(row=row_idx, column=9, value="\n".join(last_slots_text))
        
        # 模板内容
        ws.cell(row=row_idx, column=10, value=template.get("content", ""))
        
        row_idx += 1
    
    # 设置列宽
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        # 设置一个合适的列宽
        if col_idx in [6, 9, 10]:  # 槽位和内容列宽度大一些
            ws.column_dimensions[column_letter].width = 40
        else:
            ws.column_dimensions[column_letter].width = 20
    
    # 设置所有数据单元格自动换行
    for row in ws.iter_rows(min_row=2, max_row=row_idx-1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # 保存Excel文件
    try:
        wb.save(excel_file_path)
        print(f"Excel文件已保存至: {excel_file_path}")
        return True
    except Exception as e:
        print(f"保存Excel文件失败: {e}")
        return False

# 主函数
def main():
    # 获取当前脚本所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    json_file_path = os.path.join(current_dir, "template.json")
    excel_file_path = os.path.join(current_dir, "template.xlsx")
    
    # 转换JSON到Excel
    success = json_to_excel(json_file_path, excel_file_path)
    
    if success:
        print("转换完成！")
    else:
        print("转换失败！")

if __name__ == "__main__":
    main()